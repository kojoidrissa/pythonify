from openpyxl import Workbook
from openpyxl import load_workbook
import time #only here for profiling purposes

#Open the workbook, get the sheet in question
time0 = time.time()
source = load_workbook('testKojo_Atlantic Engineering CC Budgets for 2012_2017_JunQtr Update_9 16 13_Aug Actuals updated.xlsx')
time1 = time.time()
print "Time to load source worksheet", time1 - time0, "seconds."

#Create target workbook for writing
target = Workbook()
target_filename = r'testKojo_Atlantic Engineering CC Budgets Modified Dupe.xlsx'

allsheets = source.get_sheet_names() #get list of sheet names

#duplicate source into target workbook
for s in allsheets:
	target.add_sheet(source.get_sheet_by_name(s))

#Modify cells in target workbook
datasheets = allsheets[1:] 
time0 = time.time()
for s in datasheets:
	# ws = source.get_sheet_by_name(s)
	# maxrow = ws.get_highest_row()
	# maxcol = ws.get_highest_column()
	# print "Sheet", s, "has", maxrow, "rows and", maxcol, "columns"
	
	# wst = target.create_sheet()
	print 
	target.get_sheet_by_name(s).cell('F1').value = 'LEMON'


	# for r in range(1, (maxrow + 1)):
	#     for c in range(1, (maxcol + 1):
	    

time1 = time.time()
print "Time to count rows & columns in all sheets:", time1 - time0, "seconds."

#commented this out of "functionalSheets.py" 
#but THIS might work for my exceptions.py code; will .value let me copy/capture formulas?
'''
for c in r:
    ci = headcount_sorted[ri].index(c)
    #print "Cell #, Value:", ci, c
    ws.cell(row = ri, column = ci).value = headcount_sorted[ri][ci]
'''   


#write target workbook
target.save(target_filename)